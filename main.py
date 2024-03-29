from tkinter import messagebox
from openpyxl import Workbook
from tkinter.ttk import Combobox
import sqlite3 as sq
from openpyxl import load_workbook
import random
import string
from tkinter import *
from tkinter import filedialog as fd
from docx import Document
import win32com.client as win32
from docx.shared import Inches
from out_form import *




# создание окна
ws = Tk()
ws.title('Ввод значений')
ws.geometry('800x620')
ws["bg"] = "gray80"

# создание базы
conn = sq.connect('Пациенты1.db')
cur = conn.cursor()


def generate_key(length):
    letters_and_digits = string.ascii_letters + string.digits
    rand_string = ''.join(random.sample(letters_and_digits, length))
    return rand_string

def Count():
    cur.execute("SELECT num FROM анализы")
    count_row = len(cur.fetchall())
    conn.commit()
    return str(count_row)

def Input():
    key = New()
    data_list = [(Count(), key, str(Surname.get()), str(name_ManWom.get()), str(name_dateAnal.get()), str(name_age.get()), str(name_DiagnosBasic.get()), str(name_DiagnosRelated.get()), str(name_gens1.get()), str(name_2gens.get()), str(name_3gens.get()), str(name_Season.get()),str(name_WBC.get()), str(name_LYMF.get()), str(name_MON.get()), str(name_NEU.get()), str(name_HGB.get()),str(name_PLT.get()), str(name_EOS.get()), str(name_BAS.get()), str(name_CD3.get()), str(name_CD19.get()), str(name_CD4.get()), str(name_CD8.get()), str(name_ratio.get()), str(name_NK_Ob.get()), str(name_NK_Cit.get()), str(name_CIK.get()), str(name_HCI_CO.get()), str(name_HCI_CT.get()), str(name_CD3IFN_CO.get()), str(name_CD3IFN_CT.get()), str(name_CD3INF_CO.get()),str(name_CD3INF_CT.get()), str(name_CD3IL2_CO.get()), str(name_CD3IL2_CT.get()), str(name_FNO.get()))]
    cur.executemany("INSERT INTO анализы VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);", data_list)
    conn.commit()

    Graphics()
    SaveGraphicInPNG(Count())


    return messagebox.showinfo('Вывод', "Данные сохраненны")

def Graphics():
    # конвертация введенных данных
    lymf = float(name_LYMF.get())
    neu = float(name_NEU.get())
    cd3 = float(name_CD3.get())
    cd4 = float(name_CD4.get())
    cd8 = float(name_CD8.get())
    cd19 = float(name_CD19.get())

    lymf = float(name_LYMF.get())
    neu = float(name_NEU.get())
    cd3 = float(name_CD3.get())
    cd4 = float(name_CD4.get())
    cd8 = float(name_CD8.get())
    cd19 = float(name_CD19.get())
    date = str(name_dateAnal.get())

    CD3IFN_CO =  float(name_CD3IFN_CO.get())
    CD3IFN_CT =  float(name_CD3IFN_CT.get())

    CD3INF_CT =  float(name_CD3INF_CT.get())
    CD3INF_CO =  float(name_CD3IFN_CO.get())

    CD3IL2_CO =  float(name_CD3IL2_CO.get())
    CD3IL2_CT =  float(name_CD3IL2_CT.get())

    interferon = CD3IFN_CT/ CD3IFN_CO
    fno =CD3INF_CT/CD3INF_CO
    interlikin =CD3IL2_CT/CD3IL2_CO

    # создание графиков
    wb = Workbook()
    ws = wb.active
    curr_wb = load_workbook('Графики.xlsx')
    ws = curr_wb['grafik']

    ws['F14'] = neu / lymf
    ws['F15'] = neu / cd3
    ws['F16'] = neu / cd4
    ws['F17'] = neu / cd8

    ws['F23'] = neu / lymf
    ws['F24'] = lymf / cd19
    ws['F25'] = cd19 / cd4
    ws['F26'] = cd19 / cd8

    ws['M22'] = interferon
    ws['M23'] = fno
    ws['M24'] = interlikin

    number = Count()
    curr_wb.save(f'{number}results.xlsx')
    curr_wb.close()

def Graphics2():
    lymf = float(name_LYMF.get())
    neu = float(name_NEU.get())
    cd3 = float(name_CD3.get())
    cd4 = float(name_CD4.get())
    cd8 = float(name_CD8.get())
    cd19 = float(name_CD19.get())
    date = str(name_dateAnal.get())

    wb2 = Workbook()
    ws2 = wb2.active
    curr_wb2 = load_workbook('ГрафикиСезоны.xlsx')
    ws2 = curr_wb2['гипотеза']

    ws2['C3'] = date
    ws2['C4'] = cd8
    ws2['C7'] = cd4
    ws2['C10'] = neu / lymf
    ws2['C13'] = neu / cd3
    ws2['C16'] = neu / cd8
    ws2['C19'] = neu / cd4

    ws2['C25'] = date
    ws2['C26'] = cd8
    ws2['C29'] = cd4
    ws2['C32'] = neu / lymf
    ws2['C35'] = neu / cd19
    ws2['C38'] = neu / cd4
    ws2['C41'] = neu / cd8

    number = Count()
    curr_wb2.save(f'{number}grapic.xlsx')
    curr_wb2.close()

def Graphics3():
    # конвертация введенных данных
    cd3 = float(name_CD3.get())
    cd4 = float(name_CD4.get())
    cd8 = float(name_CD8.get())
    surname = str(Surname.get())

    numb = Count()
    # создание графиков
    wb = Workbook()
    ws = wb.active
    curr_wb = load_workbook('graphic3.xlsx')
    ws = curr_wb['grapic']
    ws[f'A{numb}'] = (cd3+cd4)/2
    ws[f'B{numb}'] = surname

    ws[f'E{numb}'] = (cd3+cd8)/2
    ws[f'D{numb}'] = surname

    curr_wb.save('graphic3.xlsx')
    curr_wb.close()

def Delete():
    Surname.delete("0", END)
    name_ManWom.delete("0", END)
    name_DiagnosBasic.delete("0", END)
    name_DiagnosRelated.delete("0", END)
    name_Season.delete("0", END)
    name_dateAnal.delete("0", END)
    name_LYMF.delete("0", END)
    name_NEU.delete("0", END)
    name_CD3.delete("0", END)
    name_CD4.delete("0", END)
    name_CD8.delete("0", END)
    name_CD19.delete("0", END)
    name_WBC.delete("0", END)
    name_CD3IFN_CT.delete("0", END)
    name_CD3IFN_CO.delete("0", END)
    name_MON.delete("0", END)
    name_CD3INF_CT.delete("0", END)
    name_age.delete("0", END)
    name_CD3INF_CO.delete("0", END)
    name_EOS.delete("0", END)
    name_ratio.delete("0", END)
    name_CD3IL2_CT.delete("0", END)
    name_BAS.delete("0", END)
    name_NK_Ob.delete("0", END)
    name_CD3IL2_CO.delete("0", END)
    name_gens1.delete("0", END)
    name_HGB.delete("0", END)
    name_NK_Cit.delete("0", END)
    name_FNO.delete("0", END)
    name_2gens.delete("0", END)
    name_PLT.delete("0", END)
    name_CIK.delete("0", END)
    name_3gens.delete("0", END)
    name_HCI_CO.delete("0", END)
    name_HCI_CT.delete("0", END)

def New():
    key = generate_key(5)
    return key

def insert_text():
    file_name = fd.askopenfilename()
    f = open(file_name)
    s = f.read()
    f.close()

#создание объектов
Label(ws, bg='gray81', font="TkDefaultFont 11",text="Выберите из списка или создайте новую").place(x=20, y=10)

combo = Combobox(ws)
cur.execute("SELECT фамилия FROM ключи")

combo['values'] = cur.fetchall()
combo.current(1)
combo.place(x=22, y=35)

btn_Delete = Button(ws, text="    НОВЫЙ    ", command=New)
btn_Delete.place(x=200, y=33)

btn_Delete = Button(ws, text="    Cоздание файла    ", command=createWordFile)
btn_Delete.place(x=300, y=33)

Label(ws, bg='gray81', font="TkDefaultFont 11",text="Или выберите docx файл").place(x=475, y=10)
btn_Parser = Button(ws, text="    Выбрать файл    ", command=insert_text)
btn_Parser.place(x=500, y=33)

btn_Input = Button(ws, text="    Ввод    ", command=Input)
btn_Input.place(x=700, y=20)

#0 заглавки

Label(ws, bg='gray81', font="TkDefaultFont 11",text="Персональные данные").place(x=25, y=60)
Label(ws, bg='gray81',font="TkDefaultFont 11", text="Резульаты гемоталогического\nисследования").place(x=205, y=60)
Label(ws, bg='gray81',font="TkDefaultFont 11", text="Имунный статус").place(x=440, y=63)
Label(ws, bg='gray81',font="TkDefaultFont 11", text="Цитокиновый статус").place(x=625, y=62)

#1

Label(ws, bg='gray81', text="Фамилия").place(x=70, y=100)
Surname = Entry(ws)
Surname.place(x=37.5, y=120)

Label(ws, bg='gray81', text="Лейкоциты(WBC)").place(x=252, y=100)
name_WBC = Entry(ws)
name_WBC.place(x=237.5, y=120)

Label(ws, bg='gray81', text="Общие Т-Лимфоциты").place(x=440, y=100)
name_CD3 = Entry(ws)
name_CD3.place(x=437.5, y=120)

Label(ws, bg='gray81', text="CD3+IFN+(стимулированный)").place(x=620, y=100)
name_CD3IFN_CT = Entry(ws)
name_CD3IFN_CT.place(x=637.5, y=120)

#2

Label(ws, bg='gray81', text="Пол(м/ж)").place(x=70, y=150)
name_ManWom = Entry(ws)
name_ManWom.place(x=37.5, y=170)

Label(ws, bg='gray81', text="Леймфоциты(LYMF)").place(x=247, y=150)
name_LYMF = Entry(ws)
name_LYMF.place(x=237.5, y=170)

Label(ws, bg='gray81', text="Общие B-Лимфоциты").place(x=440, y=150)
name_CD19 = Entry(ws)
name_CD19.place(x=437.5, y=170)

Label(ws, bg='gray81', text="CD3+IFN+(спонтанный)").place(x=635, y=150)
name_CD3IFN_CO = Entry(ws)
name_CD3IFN_CO.place(x=637.5, y=170)

#3

Label(ws, bg='gray81', text="Дата анализа").place(x=67, y=200)
name_dateAnal = Entry(ws)
name_dateAnal.place(x=37.5, y=220)

Label(ws, bg='gray81', text="Моноциты(MON)").place(x=252, y=200)
name_MON = Entry(ws)
name_MON.place(x=237.5, y=220)

Label(ws, bg='gray81', text="T-хелперы").place(x=470, y=200)
name_CD4 = Entry(ws)
name_CD4.place(x=437.5, y=220)

Label(ws, bg='gray81', text="CD3+TNFa+(стимулированный)").place(x=610, y=200)
name_CD3INF_CT = Entry(ws)
name_CD3INF_CT.place(x=637.5, y=220)

#4

Label(ws, bg='gray81', text="Возраст пациента").place(x=52, y=250)
name_age = Entry(ws)
name_age.place(x=37.5, y=270)

Label(ws, bg='gray81', text="Нейтрофилы(NEU)").place(x=247, y=250)
name_NEU = Entry(ws)
name_NEU.place(x=237.5, y=270)

Label(ws, bg='gray81', text="Т-цитотоксические лимфоциты").place(x=415, y=250)
name_CD8 = Entry(ws)
name_CD8.place(x=437.5, y=270)

Label(ws, bg='gray81', text="CD3+TNFa+(спонтанный)").place(x=625, y=250)
name_CD3INF_CO = Entry(ws)
name_CD3INF_CO.place(x=637.5, y=270)

#5

Label(ws, bg='gray81', text="Диагноз оснвной").place(x=51, y=300)
name_DiagnosBasic = Entry(ws)
name_DiagnosBasic.place(x=37.5, y=320)

Label(ws, bg='gray81', text="Эозофилы(EOS)").place(x=257, y=300)
name_EOS = Entry(ws)
name_EOS.place(x=237.5, y=320)

Label(ws, bg='gray81', text="Соотношение CD3+CD4+/CD3+CD8+").place(x=405, y=300)
name_ratio = Entry(ws)
name_ratio.place(x=437.5, y=320)

Label(ws, bg='gray81', text="CD3+IL2+(стимулированный)").place(x=620, y=300)
name_CD3IL2_CT = Entry(ws)
name_CD3IL2_CT.place(x=637.5, y=320)

#6

Label(ws, bg='gray81', text="Дигноз сопустствующий").place(x=32, y=350)
name_DiagnosRelated = Entry(ws)
name_DiagnosRelated.place(x=37.5, y=370)

Label(ws, bg='gray81', text="Базофилы(BAS)").place(x=257, y=350)
name_BAS = Entry(ws)
name_BAS.place(x=237.5, y=370)

Label(ws, bg='gray81', text="Общие NK-клетки").place(x=447, y=350)
name_NK_Ob = Entry(ws)
name_NK_Ob.place(x=437.5, y=370)

Label(ws, bg='gray81', text="CD3+IL2+(спонтанный)").place(x=635, y=350)
name_CD3IL2_CO = Entry(ws)
name_CD3IL2_CO.place(x=637.5, y=370)

#7

Label(ws, bg='gray81', text="1.Гены").place(x=80, y=400)
name_gens1 = Entry(ws)
name_gens1.place(x=37.5, y=420)

Label(ws, bg='gray81', text="Гемоглобин(HGB)").place(x=252, y=400)
name_HGB = Entry(ws)
name_HGB.place(x=237.5, y=420)

Label(ws, bg='gray81', text="NK-клетки цитолитические").place(x=425, y=400)
name_NK_Cit = Entry(ws)
name_NK_Cit.place(x=437.5, y=420)

Label(ws, bg='gray81', text="ФНО").place(x=685, y=400)
name_FNO = Entry(ws)
name_FNO.place(x=637.5, y=420)

#8

Label(ws, bg='gray81', text="2.Гены").place(x=80, y=450)
name_2gens = Entry(ws)
name_2gens.place(x=37.5, y=470)

Label(ws, bg='gray81', text="Тромбоциты(PLT)").place(x=252, y=450)
name_PLT = Entry(ws)
name_PLT.place(x=237.5, y=470)

Label(ws, bg='gray81', text="Цируклирующие имунные комплексы").place(x=395, y=450)
name_CIK = Entry(ws)
name_CIK.place(x=437.5, y=470)

#9

Label(ws, bg='gray81', text="3.Гены").place(x=80, y=500)
name_3gens = Entry(ws)
name_3gens.place(x=37.5, y=520)

Label(ws, bg='gray81', text="HCI-тест(спонтанный)").place(x=440, y=500)
name_HCI_CO = Entry(ws)
name_HCI_CO.place(x=437.5, y=520)

#10

Label(ws, bg='gray81', text="Сезон").place(x=82, y=550)
name_Season = Entry(ws)
name_Season.place(x=37.5, y=570)

Label(ws, bg='gray81', text="HCI-тест(стимулированный)").place(x=422, y=550)
name_HCI_CT = Entry(ws)
name_HCI_CT.place(x=437.5, y=570)

ws.mainloop()
wb = Workbook()
ws = wb.active